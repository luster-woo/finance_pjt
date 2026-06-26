import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand


DATA_DIR = Path(__file__).resolve().parents[4] / 'data'


class Command(BaseCommand):
    help = '카드(card_data.json)와 원자재(Gold/Silver .xlsx) 데이터를 DB에 삽입합니다.'

    def handle(self, *args, **options):
        self.stdout.write('=== 데이터 삽입 시작 ===\n')
        self._load_cards()
        self._load_commodities()
        self.stdout.write(self.style.SUCCESS('\n=== 완료 ==='))

    def _load_cards(self):
        from cards.models import Card, CardBenefit

        self.stdout.write('[1/2] 카드 데이터 삽입 중...')

        card_file = DATA_DIR / 'card_data.json'
        if not card_file.exists():
            self.stdout.write(self.style.WARNING(f'  파일 없음: {card_file}'))
            return

        with open(card_file, 'r', encoding='utf-8') as f:
            items = json.load(f)

        created = 0
        for item in items:
            card, is_new = Card.objects.get_or_create(
                card_name=item['card_name'],
                company=item['company'],
                defaults={
                    'card_type':       item['card_type'],
                    'min_performance': item.get('min_performance'),
                    'annual_fee':      item.get('annual_fee'),
                },
            )
            if is_new:
                CardBenefit.objects.bulk_create([
                    CardBenefit(
                        card=card,
                        benefit_category=b['category'],
                        benefit_detail=b['detail'],
                    )
                    for b in item.get('benefits', [])
                ])
                created += 1

        self.stdout.write(self.style.SUCCESS(f'  카드: {created}개 삽입 완료 (이미 있는 항목 스킵)'))

    def _load_commodities(self):
        from commodities.models import CommodityPrice

        self.stdout.write('[2/2] 원자재(금/은) 데이터 삽입 중...')

        try:
            from openpyxl import load_workbook
        except ImportError:
            self.stdout.write(self.style.WARNING('  openpyxl 없음 → pip install openpyxl 후 재실행'))
            return

        files = [
            ('Gold_prices.xlsx',   'Gold',   '금'),
            ('Silver_prices.xlsx', 'Silver', '은'),
        ]

        for filename, commodity_type, label in files:
            file_path = DATA_DIR / filename
            if not file_path.exists():
                self.stdout.write(self.style.WARNING(f'  파일 없음: {filename}'))
                continue

            wb = load_workbook(file_path, read_only=True, data_only=True)
            rows = wb[wb.sheetnames[0]].iter_rows(values_only=True)
            next(rows, None)  # 헤더 스킵

            created = 0
            for row in rows:
                if not row or len(row) < 2 or row[0] is None or row[1] is None:
                    continue

                date_val = row[0]
                if isinstance(date_val, datetime):
                    recorded_at = date_val
                else:
                    try:
                        recorded_at = datetime.strptime(str(date_val), '%Y-%m-%d')
                    except ValueError:
                        continue

                try:
                    price = Decimal(str(row[1]).replace(',', '').strip())
                except Exception:
                    continue

                _, is_new = CommodityPrice.objects.get_or_create(
                    commodity_type=commodity_type,
                    recorded_at=recorded_at,
                    defaults={'price': price},
                )
                if is_new:
                    created += 1

            self.stdout.write(self.style.SUCCESS(f'  {label}: {created}개 삽입 완료'))
