from decimal import Decimal, InvalidOperation
import datetime
from pathlib import Path

import yfinance as yf
from openpyxl import load_workbook
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils.timezone import make_aware

from .models import CommodityPrice
from .serializers import CommodityPriceSerializer

# Yahoo Finance 티커 (선물 기준, USD/oz)
COMMODITY_TICKERS = {
    'Gold':   'GC=F',   # 금 선물
    'Silver': 'SI=F',   # 은 선물
}


def _parse_xlsx_date(value):
    """엑셀 날짜 값 → timezone-aware datetime (Asia/Seoul)."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        naive = value.replace(tzinfo=None)
        return make_aware(naive)
    if isinstance(value, str):
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
            try:
                return make_aware(datetime.datetime.strptime(value, fmt))
            except ValueError:
                continue
    return None


def _save_commodity(commodity_type, recorded_at, price_decimal):
    _, created = CommodityPrice.objects.update_or_create(
        commodity_type=commodity_type,
        recorded_at=recorded_at,
        defaults={'price': price_decimal},
    )
    return created


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_commodities(request):
    """xlsx 파일 + yfinance 최신 데이터를 모두 가져와 저장합니다."""
    created_total = updated_total = 0

    # ── 1. xlsx 히스토리 데이터 ─────────────────────────────────────
    data_dir = Path(__file__).resolve().parents[2] / 'data'
    xlsx_files = [
        ('Gold_prices.xlsx', 'Gold'),
        ('Silver_prices.xlsx', 'Silver'),
    ]
    missing_files = []
    for filename, commodity_type in xlsx_files:
        file_path = data_dir / filename
        if not file_path.exists():
            missing_files.append(filename)
            continue
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)  # 헤더 스킵
        for row in rows:
            if not row or len(row) < 2:
                continue
            recorded_at = _parse_xlsx_date(row[0])
            if recorded_at is None or row[1] is None:
                continue
            try:
                price = Decimal(str(row[1]).replace(',', '').strip())
            except InvalidOperation:
                continue
            flag = _save_commodity(commodity_type, recorded_at, price)
            if flag:
                created_total += 1
            else:
                updated_total += 1

    # ── 2. yfinance 최신 데이터 (최근 2년) ─────────────────────────
    yf_errors = []
    for commodity_type, ticker_sym in COMMODITY_TICKERS.items():
        try:
            df = yf.Ticker(ticker_sym).history(period='2y')
        except Exception as exc:
            yf_errors.append(f'{commodity_type}: {exc}')
            continue
        if df.empty:
            continue
        for date, row in df.iterrows():
            close = row.get('Close')
            if close is None:
                continue
            try:
                price = Decimal(str(round(float(close), 4)))
            except (InvalidOperation, ValueError):
                continue
            # yfinance 반환 타임스탬프 → timezone-aware datetime
            raw_dt = date.to_pydatetime() if hasattr(date, 'to_pydatetime') else date
            if isinstance(raw_dt, datetime.datetime) and raw_dt.tzinfo is None:
                recorded_at = make_aware(raw_dt)
            elif isinstance(raw_dt, datetime.datetime):
                recorded_at = raw_dt
            else:
                recorded_at = make_aware(datetime.datetime.combine(raw_dt, datetime.time.min))
            flag = _save_commodity(commodity_type, recorded_at, price)
            if flag:
                created_total += 1
            else:
                updated_total += 1

    result = {'created': created_total, 'updated': updated_total}
    if missing_files:
        result['missing_xlsx'] = missing_files
    if yf_errors:
        result['yf_errors'] = yf_errors
    return Response(result)


@api_view(['GET'])
def commodity_list(request):
    commodity_type = request.GET.get('type')  # Gold or Silver
    qs = CommodityPrice.objects.all().order_by('-recorded_at')
    if commodity_type:
        qs = qs.filter(commodity_type=commodity_type)
    serializer = CommodityPriceSerializer(qs, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def commodity_summary(request):
    """금·은 최신 가격과 6개월 이력을 한번에 반환."""
    result = {}
    for ct in ['Gold', 'Silver']:
        latest = CommodityPrice.objects.filter(commodity_type=ct).order_by('-recorded_at').first()
        history = list(
            CommodityPrice.objects.filter(commodity_type=ct)
            .order_by('-recorded_at')[:180]
            .values('recorded_at', 'price')
        )
        history.reverse()
        result[ct] = {
            'latest_price': str(latest.price) if latest else None,
            'latest_date': latest.recorded_at.strftime('%Y-%m-%d') if latest else None,
            'history': [{'date': h['recorded_at'].strftime('%Y-%m-%d'), 'price': str(h['price'])} for h in history],
        }
    return Response(result)
